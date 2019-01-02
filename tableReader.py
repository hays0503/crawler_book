import openpyxl

"""
@package tableReader
"""


class TableInspector:
    """
        @brief Считывание и систиматизация данных из таблиц MS Excel
    """
    other_description = list()

    def __init__(self, patch_to_file, number_sheet=0):
        self.read_sheet(patch_to_file, number_sheet)

    def read_sheet(self, patch_to_file, number_sheet=0):
        """
            @brief Считывает страницу документа возвращая объект страница
        """
        wb = openpyxl.load_workbook(filename=patch_to_file,
                                    read_only=True)
        name_sheet = wb.sheetnames
        sheet = wb[name_sheet[number_sheet]]
        row_count = sheet.max_row
        for row in sheet.rows:
            turple = [row[1].value, row[4].value]
            if turple[0] is not None and str(turple[0]) != 'Товар':
                other_info = str(turple[0]).rsplit(', ', 4)
                other_info.append(turple[1])
                self.other_description.append(other_info)
        #for i in range(len(self.other_description)):
           #print(self.other_description.__getitem__(i))
        return self.other_description

    def name_book(self, row=0):
        if not self.other_description.__getitem__(row).__getitem__(0).split(' ')[0].isdigit():
            return self.other_description.__getitem__(row).__getitem__(0)
        else:
            return -1

    def publisher_book(self, row=0):
        if not self.other_description.__getitem__(row).__getitem__(1).split(' ')[0].isdigit():
            return self.other_description.__getitem__(row).__getitem__(1)
        else:
            return -1


    def release_date_book(self, row=0):
        shift_data = 2
        if len(self.other_description.__getitem__(row)) == 5:
            shift_data -= 1
        if self.other_description.__getitem__(row).__getitem__(shift_data).split(' ')[0].isdigit():
            return int(self.other_description.__getitem__(row).__getitem__(shift_data).split(' ')[0])
        else:
            return -1

    def book_binding_type(self, row=0):
        shift_data = 3
        if len(self.other_description.__getitem__(row)) == 5:
            shift_data -= 1
        if not self.other_description.__getitem__(row).__getitem__(shift_data).split(' ')[0].isdigit():
            return self.other_description.__getitem__(row).__getitem__(shift_data)
        else:
            return -1

    def number_of_pages_book(self, row=0):
        shift_data = 4
        if len(self.other_description.__getitem__(row)) == 5:
            shift_data -= 1
        if self.other_description.__getitem__(row).__getitem__(shift_data).split(' ')[0].isdigit():
            return int(self.other_description.__getitem__(row).__getitem__(shift_data).split(' ')[0])
        else:
            return -1

    def url(self, row=0):
        shift_data = 5
        if len(self.other_description.__getitem__(row)) == 5:
            shift_data -= 1
        if not self.other_description.__getitem__(row).__getitem__(shift_data).split(' ')[0].isdigit():
            return self.other_description.__getitem__(row).__getitem__(shift_data)
        else:
            return -1

"""
if __name__ == '__main__':
    table_Inspector = TableInspector(patch_to_file="prices_s2.xlsx")
    print(
        table_Inspector.name_book(1) + ':1\n' +
        table_Inspector.publisher_book(1) + ':2\n' +
        str(table_Inspector.release_date_book(1)) + ':3\n' +
        table_Inspector.book_binding_type(1) + ':4\n' +
        str(table_Inspector.number_of_pages_book(1)) + ':5\n'+
        str(table_Inspector.url(1)) + ':6\n')
"""